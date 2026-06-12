from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from aimaos.analyzers.performance_analyzer import AnalysisResult
from aimaos.recommenders.action_engine import OperatingActionIssue
from aimaos.reports.formatting import (
    COUNT_COLUMNS,
    DATE_COLUMNS,
    EXCEL_NUMBER_FORMATS,
    MONEY_COLUMNS,
    PERCENT_COLUMNS,
    RATIO_COLUMNS,
    column_description,
    column_key,
    column_label,
    format_count,
    format_dataframe_for_display,
    format_money,
    format_percent,
    format_ratio,
    localize_dataframe_for_excel,
    metric_key,
    sheet_name,
    value_label,
)


@dataclass(frozen=True)
class ReportPaths:
    markdown: Path
    text: Path
    excel: Path


def build_markdown_report(
    advertiser: str,
    analysis: AnalysisResult,
    recommendations: pd.DataFrame,
    mapping: dict[str, str],
    unmapped_columns: list[str],
    action_issues: list[OperatingActionIssue] | None = None,
) -> str:
    summary = analysis.summary
    lines = [
        f"# AIMAOS 광고 성과 분석 보고서 - {advertiser}",
        "",
        "## 1. 핵심 요약",
        "",
        f"- 분석 행 수: {format_count(summary['rows'])}",
        f"- 노출수: {format_count(summary['impressions'])}",
        f"- 클릭수: {format_count(summary['clicks'])}",
        f"- 광고비: {format_money(summary['cost'])}",
        f"- 전환수: {format_count(summary['conversions'])}",
        f"- 매출: {format_money(summary['revenue'])}",
        f"- 클릭률: {format_percent(summary['ctr'])}",
        f"- 클릭당 비용: {format_money(summary['cpc'])}",
        f"- 전환율: {format_percent(summary['cvr'])}",
        f"- 전환당 비용: {format_money(summary['cpa'])}",
        f"- 광고수익률: {format_ratio(summary['roas'])}",
        "",
        "## 2. 우선 실행 권고",
        "",
    ]

    for _, row in recommendations.iterrows():
        lines.append(
            f"- [{value_label(row['priority'])}] {row['target']}: {row['recommended_action']} "
            f"({row['expected_effect']})"
        )

    lines.extend(["", "## 3. 오늘 해야 할 일", ""])
    lines.extend(_today_action_markdown_lines(action_issues or []))

    lines.extend(["", "## 4. 운영 점검 포인트", ""])
    if analysis.anomalies.empty:
        lines.append("- 우선 확인할 운영 점검 포인트가 없습니다.")
    else:
        for _, row in analysis.anomalies.iterrows():
            lines.append(f"- {value_label(row['severity'])} / {row['type']} / {row['target']}: {row['reason']}")

    lines.extend(["", "## 5. 기간별 성과", ""])
    if analysis.period_tables:
        lines.extend(
            [
                f"- 일자별: 엑셀 `{sheet_name('daily')}` 시트",
                f"- 주간별: 엑셀 `{sheet_name('weekly')}` 시트",
                f"- 월별: 엑셀 `{sheet_name('monthly')}` 시트",
                f"- 분기별: 엑셀 `{sheet_name('quarterly')}` 시트",
                f"- 반기별: 엑셀 `{sheet_name('half_year')}` 시트",
                f"- 연도별: 엑셀 `{sheet_name('yearly')}` 시트",
                f"- 시즌별: 엑셀 `{sheet_name('seasonal')}` 시트",
                "",
            ]
        )
        monthly = analysis.period_tables.get("monthly", pd.DataFrame())
        weekly = analysis.period_tables.get("weekly", pd.DataFrame())
        if not monthly.empty:
            lines.extend(["### 월별 요약", "", _markdown_table(monthly), ""])
        if not weekly.empty:
            lines.extend(["### 주간별 요약", "", _markdown_table(weekly), ""])
    else:
        lines.append("- 날짜 컬럼이 없어 기간별 성과를 생성하지 못했습니다.")

    lines.extend(["", "## 6. 컬럼 매핑 결과", ""])
    for source, target in mapping.items():
        lines.append(f"- {source} -> {column_label(target)}")

    if unmapped_columns:
        lines.extend(["", "### 미매핑 컬럼", ""])
        for column in unmapped_columns:
            lines.append(f"- {column}")

    lines.extend(["", "## 7. 운영 메모", ""])
    lines.append("- 본 보고서는 자동 분석 초안입니다. 긴급 항목은 운영자가 원본 광고센터에서 직접 확인한 뒤 실행합니다.")
    lines.append("- 분석 결과와 실행 여부는 지식 자산화를 위해 사례 DB에 누적하는 것을 권장합니다.")

    return "\n".join(lines) + "\n"


def _today_action_markdown_lines(action_issues: list[OperatingActionIssue]) -> list[str]:
    if not action_issues:
        return ["- 현재 Rule Engine 기준 주요 조치 필요 항목 없음"]

    major_issues = [
        issue
        for issue in action_issues
        if issue.report_include or issue.problem in {"판단 불가", "데이터 부족"} or issue.severity in {"high", "medium", "opportunity"}
    ]
    if not major_issues:
        return ["- 현재 Rule Engine 기준 주요 조치 필요 항목 없음"]

    lines: list[str] = []
    for index, issue in enumerate(major_issues, start=1):
        lines.extend(
            [
                f"### {index}. [{_severity_label(issue.severity)}] {issue.problem}",
                "",
                f"- 심각도: {_severity_label(issue.severity)}",
                f"- 우선순위점수: {issue.priority_score:.2f}",
                f"- 광고주명: {issue.advertiser}",
                f"- 매체: {issue.media}",
                f"- 대상: {issue.target}",
                f"- 근거데이터: {issue.evidence}",
                f"- 원인추정: {issue.cause_hypothesis}",
                f"- 추천액션: {issue.recommended_action}",
                f"- 예상효과: {issue.expected_effect}",
                f"- 광고주설명문구: {issue.advertiser_message}",
                f"- 보고서반영여부: {'반영 권장' if issue.report_include else '필요 시 반영'}",
                "",
            ]
        )
    return lines


def _markdown_table(frame: pd.DataFrame) -> str:
    columns = [
        "period_label",
        "period_start",
        "period_end",
        "impressions",
        "clicks",
        "cost",
        "conversions",
        "revenue",
        "ctr",
        "cpc",
        "cvr",
        "cpa",
        "roas",
    ]
    available = [column for column in columns if column in frame.columns]
    displayed = format_dataframe_for_display(frame[available])
    headers = list(displayed.columns)
    rows = displayed.astype(str).values.tolist()
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def _today_actions_dataframe(action_issues: list[OperatingActionIssue]) -> pd.DataFrame:
    columns = [
        "심각도",
        "우선순위점수",
        "광고주명",
        "매체",
        "대상",
        "발견된문제",
        "근거데이터",
        "원인추정",
        "추천액션",
        "예상효과",
        "광고주설명문구",
        "보고서반영여부",
    ]
    rows = [
        {
            "심각도": _severity_label(issue.severity),
            "우선순위점수": issue.priority_score,
            "광고주명": issue.advertiser,
            "매체": issue.media,
            "대상": issue.target,
            "발견된문제": issue.problem,
            "근거데이터": issue.evidence,
            "원인추정": issue.cause_hypothesis,
            "추천액션": issue.recommended_action,
            "예상효과": issue.expected_effect,
            "광고주설명문구": issue.advertiser_message,
            "보고서반영여부": "반영 권장" if issue.report_include else "필요 시 반영",
        }
        for issue in action_issues
    ]
    return pd.DataFrame(rows, columns=columns)


def _severity_label(severity: str) -> str:
    return {
        "high": "긴급",
        "medium": "관찰",
        "opportunity": "기회",
        "low": "참고",
    }.get(severity, "점검")


def write_reports(
    output_dir: str | Path,
    advertiser: str,
    standardized: pd.DataFrame,
    analysis: AnalysisResult,
    recommendations: pd.DataFrame,
    mapping: dict[str, str],
    unmapped_columns: list[str],
    action_issues: list[OperatingActionIssue] | None = None,
) -> ReportPaths:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    action_issues = action_issues or []
    markdown = build_markdown_report(advertiser, analysis, recommendations, mapping, unmapped_columns, action_issues)

    markdown_path = out / "report.md"
    text_path = out / "report.txt"
    excel_path = out / "analysis.xlsx"

    markdown_path.write_text(markdown, encoding="utf-8")
    text_path.write_text(markdown.replace("#", "").replace("*", ""), encoding="utf-8")

    try:
        _write_excel_report(excel_path, standardized, analysis, recommendations, action_issues)
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = out / f"analysis_{timestamp}.xlsx"
        _write_excel_report(excel_path, standardized, analysis, recommendations, action_issues)

    return ReportPaths(markdown=markdown_path, text=text_path, excel=excel_path)


def _write_excel_report(
    excel_path: Path,
    standardized: pd.DataFrame,
    analysis: AnalysisResult,
    recommendations: pd.DataFrame,
    action_issues: list[OperatingActionIssue],
) -> None:
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        _build_guide_sheet().to_excel(writer, index=False, sheet_name=sheet_name("guide"))
        localize_dataframe_for_excel(standardized).to_excel(writer, index=False, sheet_name=sheet_name("standardized"))
        localize_dataframe_for_excel(pd.DataFrame([analysis.summary])).to_excel(
            writer,
            index=False,
            sheet_name=sheet_name("summary"),
        )
        localize_dataframe_for_excel(recommendations).to_excel(
            writer,
            index=False,
            sheet_name=sheet_name("recommendations"),
        )
        _today_actions_dataframe(action_issues).to_excel(
            writer,
            index=False,
            sheet_name=sheet_name("today_actions"),
        )
        localize_dataframe_for_excel(analysis.anomalies).to_excel(
            writer,
            index=False,
            sheet_name=sheet_name("anomalies"),
        )
        localize_dataframe_for_excel(analysis.period_comparison).to_excel(
            writer,
            index=False,
            sheet_name=sheet_name("period_compare"),
        )
        for name, table in analysis.segment_tables.items():
            if not table.empty:
                localize_dataframe_for_excel(table).to_excel(writer, index=False, sheet_name=sheet_name(name))
        for name, table in analysis.period_tables.items():
            if not table.empty:
                localize_dataframe_for_excel(table).to_excel(writer, index=False, sheet_name=sheet_name(name))

        _format_workbook(writer.book)


def _build_guide_sheet() -> pd.DataFrame:
    rows = [
        {
            "항목": "파일 사용법",
            "설명": "상단 첫 행은 항목명입니다. 각 항목명 셀에 마우스를 올리면 설명 메모를 볼 수 있습니다.",
            "계산식/활용": "먼저 전체요약, 실행권고, 운영점검포인트 시트를 보고 세부 시트를 확인하세요.",
        },
        {
            "항목": "틀고정",
            "설명": "모든 시트는 첫 행과 첫 열을 고정했습니다. 아래나 오른쪽으로 이동해도 항목명을 계속 볼 수 있습니다.",
            "계산식/활용": "데이터가 많아질수록 항목명과 기간/대상을 놓치지 않게 하기 위한 기본 설정입니다.",
        },
        {
            "항목": "숫자 표시",
            "설명": "노출수, 클릭수, 광고비, 매출은 1,000 단위 쉼표로 표시합니다.",
            "계산식/활용": "엑셀 안의 값은 계산 가능한 숫자로 유지됩니다.",
        },
        {
            "항목": "비율 표시",
            "설명": "클릭률, 전환율, 증감률, 광고수익률은 0.00% 형식으로 표시합니다.",
            "계산식/활용": "예: 클릭률 0.0141은 1.41%, 광고수익률 4.15는 415.00%로 표시됩니다.",
        },
        {
            "항목": "기간별 분석",
            "설명": "일자별, 주간별, 월별, 분기별, 반기별, 연도별, 시즌별 성과 시트를 제공합니다.",
            "계산식/활용": "월별성과와 주간별성과로 흐름을 본 뒤 일자별성과에서 원인을 확인하세요.",
        },
        {
            "항목": "오늘 해야 할 일",
            "설명": "Rule Engine이 기존 KPI와 운영 점검 결과를 참조해 생성한 운영 액션 목록입니다.",
            "계산식/활용": "오늘해야할일 시트에서 광고주에게 설명할 문구와 추천 액션을 먼저 확인하세요.",
        },
    ]
    metric_rows = [
        ("노출수", column_description("impressions"), "광고가 얼마나 많이 보였는지 확인합니다."),
        ("클릭수", column_description("clicks"), "광고가 얼마나 많은 방문을 만들었는지 확인합니다."),
        ("광고비", column_description("cost"), "광고 운영 비용 규모를 확인합니다."),
        ("전환수", column_description("conversions"), "문의, 구매, 신청 등 실제 성과 횟수를 확인합니다."),
        ("매출", column_description("revenue"), "광고로 발생한 매출 또는 전환 가치를 확인합니다."),
        ("클릭률", column_description("ctr"), "소재나 키워드가 클릭을 잘 유도하는지 판단합니다."),
        ("클릭당 비용", column_description("cpc"), "클릭을 얻는 비용이 비싼지 싼지 판단합니다."),
        ("전환율", column_description("cvr"), "클릭 이후 성과로 이어지는 힘을 판단합니다."),
        ("전환당 비용", column_description("cpa"), "전환 1건을 만드는 비용이 적절한지 판단합니다."),
        ("광고수익률", column_description("roas"), "100.00%는 광고비와 같은 금액의 매출이 발생했다는 뜻입니다."),
    ]
    rows.extend({"항목": item, "설명": description, "계산식/활용": usage} for item, description, usage in metric_rows)
    return pd.DataFrame(rows)


def _format_workbook(workbook) -> None:  # noqa: ANN001
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for worksheet in workbook.worksheets:
        headers = [cell.value for cell in worksheet[1]]
        header_by_index = {index + 1: value for index, value in enumerate(headers)}
        worksheet.freeze_panes = "B2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = True

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            description = _header_description(cell.value)
            if description:
                cell.comment = Comment(description, "AIMAOS")

        for column_cells in worksheet.columns:
            header = column_cells[0].value
            max_text_length = max(len(str(cell.value or "")) for cell in column_cells[:50])
            width = max(12, min(42, max(len(str(header or "")), max_text_length) + 4))
            worksheet.column_dimensions[column_cells[0].column_letter].width = width

        for row in worksheet.iter_rows(min_row=2):
            row_metric = _metric_from_row(headers, row)
            for cell in row:
                header = header_by_index.get(cell.column)
                cell.number_format = _excel_number_format(header, row_metric)
                cell.alignment = Alignment(vertical="center")


def _metric_from_row(headers: list[str], row) -> str | None:  # noqa: ANN001
    metric_header = "지표" if "지표" in headers else "metric"
    if metric_header not in headers:
        return None
    metric_index = headers.index(metric_header)
    return metric_key(row[metric_index].value)


def _excel_number_format(header: str | None, row_metric: str | None = None) -> str:
    key = column_key(header)
    if key in DATE_COLUMNS:
        return EXCEL_NUMBER_FORMATS["date"]
    if key in COUNT_COLUMNS:
        return EXCEL_NUMBER_FORMATS["count"]
    if key in MONEY_COLUMNS:
        return EXCEL_NUMBER_FORMATS["money"]
    if key in PERCENT_COLUMNS:
        return EXCEL_NUMBER_FORMATS["percent"]
    if key in RATIO_COLUMNS:
        return EXCEL_NUMBER_FORMATS["ratio"]
    if key in {"previous", "current", "delta"} and row_metric:
        if row_metric in COUNT_COLUMNS:
            return EXCEL_NUMBER_FORMATS["count"]
        if row_metric in MONEY_COLUMNS:
            return EXCEL_NUMBER_FORMATS["money"]
        if row_metric in PERCENT_COLUMNS:
            return EXCEL_NUMBER_FORMATS["percent"]
        if row_metric in RATIO_COLUMNS:
            return EXCEL_NUMBER_FORMATS["ratio"]
    return "General"


def _header_description(header: str | None) -> str:
    manual = {
        "항목": "설명을 제공하는 대상 항목입니다.",
        "설명": "초보자도 이해할 수 있도록 적은 쉬운 설명입니다.",
        "계산식/활용": "해당 항목을 어떻게 계산하거나 실무에서 어떻게 보면 좋은지에 대한 안내입니다.",
        "심각도": "Rule Engine이 분류한 조치 필요 수준입니다. 긴급, 관찰, 기회, 참고 등으로 표시합니다.",
        "우선순위점수": "광고비 영향도와 심각도 등을 반영한 정렬용 점수입니다.",
        "광고주명": "분석 대상 광고주 또는 계정 이름입니다.",
        "발견된문제": "Rule Engine이 찾은 운영상 확인 필요 항목입니다.",
        "근거데이터": "이슈 판단에 사용한 주요 KPI 값입니다.",
        "원인추정": "단정이 아닌 가능성 또는 검토 필요 수준의 원인 설명입니다.",
        "추천액션": "운영자가 확인하거나 실행할 수 있는 다음 행동입니다.",
        "예상효과": "추천 액션을 검토했을 때 기대할 수 있는 효과 범위입니다.",
        "광고주설명문구": "광고주에게 설명할 때 사용할 수 있는 초안 문구입니다.",
        "보고서반영여부": "월간 또는 광고주 보고서에 포함할지에 대한 권장 상태입니다.",
    }
    if header in manual:
        return manual[header]
    return column_description(header)
